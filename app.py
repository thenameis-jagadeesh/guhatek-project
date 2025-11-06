from flask import Flask, jsonify, request, render_template, redirect, url_for, session
from flask_cors import CORS
import os
import openpyxl
from datetime import datetime
import random
import json
import secrets
import sqlite3
import hashlib

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)  # Generate a secure secret key
CORS(app)

EXCEL_FILE = 'data.xlsx'
SHEET_NAME = 'Candidates'
USER_DB = 'instance/users.db'

# Default admin credentials
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "password123"

# Create sample Excel file if it doesn't exist
def create_sample_excel():
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = SHEET_NAME
    
    # Define headers
    headers = [
        'Date', 'Name', 'Email ID', 'Contact Number', 'Interested Position', 
        'Current Role', 'Current Organization', 'Current Location',
        'Current CTC per Annum', 'Expected CTC per Annum', 'Total Years of Experience',
        'Notice Period', 'In Notice', 'Immediate Joiner', 'Offers in Hand',
        'Offered CTC', 'Location Preference', 'Certifications', 'Resume',
        'LinkedIn Profile', 'Comments', 'Referred By', 'Interview Status',
        'Application Status',
        # Stage-specific remarks
        'Initial Screening', 'Round 1 Remarks', 'Round 2 Remarks',
        # General remarks
        'Remarks', 'Reject Mail Sent', 'Final Remarks',
        'Month Count', 'Reference'
    ]
    
    # Add headers to the first row
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header
        
        # Sample data
        sample_data = [
            {
                'Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Name': 'John Doe',
                'Email ID': 'john.doe@example.com',
                'Contact Number': '9876543210',
                'Interested Position': 'Software Developer',
                'Current Role': 'Junior Developer',
                'Current Organization': 'Tech Solutions Inc.',
                'Current Location': 'Bangalore',
                'Current CTC per Annum': '800000',
                'Expected CTC per Annum': '1200000',
                'Total Years of Experience': '2-3 years',
                'Notice Period': '30 days',
                'In Notice': 'Yes',
                'Immediate Joiner': 'No',
                'Offers in Hand': 'No',
                'Offered CTC': '',
                'Location Preference': 'Bangalore',
                'Certifications': 'AWS Certified Developer',
                'Resume': 'https://example.com/resume/johndoe',
                'LinkedIn Profile': 'https://linkedin.com/in/johndoe',
                'Comments': 'Good communication skills',
                'Referred By': 'Employee Referral',
                'Interview Status': 'Scheduled',
                'Application Status': 'In Process',
                'Remarks': 'Promising candidate',
                'Reject Mail Sent': 'No',
                'Initial Screening': 'Candidate performed well in initial screening.',
                'Round 1 Remarks': 'Strong technical skills demonstrated in Round 1.',
                'Round 2 Remarks': 'Good problem-solving approach in Round 2.',
                'Final Remarks': '',
                'Month Count': '1',
                'Reference': 'Jane Smith'
            },
            {
                'Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Name': 'Jane Smith',
                'Email ID': 'jane.smith@example.com',
                'Contact Number': '8765432109',
                'Interested Position': 'Data Scientist',
                'Current Role': 'Data Analyst',
                'Current Organization': 'Data Insights Ltd.',
                'Current Location': 'Hyderabad',
                'Current CTC per Annum': '1000000',
                'Expected CTC per Annum': '1500000',
                'Total Years of Experience': '3-5 years',
                'Notice Period': '60 days',
                'In Notice': 'No',
                'Immediate Joiner': 'No',
                'Offers in Hand': 'Yes',
                'Offered CTC': '1400000',
                'Location Preference': 'Remote',
                'Certifications': 'Google Data Analytics',
                'Resume': 'https://example.com/resume/janesmith',
                'LinkedIn Profile': 'https://linkedin.com/in/janesmith',
                'Comments': 'Strong analytical skills',
                'Referred By': 'Job Portal',
                'Interview Status': 'Selected',
                'Application Status': 'Offer Made',
                'Remarks': 'Top candidate',
                'Reject Mail Sent': 'No',
                'Initial Remarks': '',
                'Round 1 Remarks': '',
                'Round 2 Remarks': '',
                'Final Remarks': 'Waiting for candidate response',
                'Month Count': '2',
                'Reference': 'Robert Johnson'
            },
            {
                'Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Email ID': 'sam.wilson@example.com',
                'Contact Number': '7654321098',
                'Interested Position': 'UI/UX Designer',
                'Current Role': 'Graphic Designer',
                'Current Organization': 'Creative Designs',
                'Current Location': 'Chennai',
                'Current CTC per Annum': '700000',
                'Expected CTC per Annum': '1000000',
                'Total Years of Experience': '1-2 years',
                'Notice Period': '15 days',
                'In Notice': 'Yes',
                'Immediate Joiner': 'Yes',
                'Offers in Hand': 'No',
                'Offered CTC': '',
                'Location Preference': 'Chennai',
                'Certifications': 'Adobe Certified Expert',
                'Resume': 'https://example.com/resume/samwilson',
                'LinkedIn Profile': 'https://linkedin.com/in/samwilson',
                'Comments': 'Creative portfolio',
                'Referred By': 'Campus Recruitment',
                'Interview Status': 'Rejected',
                'Application Status': 'Rejected',
                'Remarks': 'Not enough experience',
                'Reject Mail Sent': 'Yes',
                'Initial Remarks': '',
                'Round 1 Remarks': '',
                'Round 2 Remarks': '',
                'Final Remarks': 'Consider for junior positions',
                'Month Count': '1',
                'Reference': 'Emily Davis'
            }
        ]
        
        # Add sample data
        for row_num, data in enumerate(sample_data, 2):
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num).value = data.get(header, '')
        
        # Save the workbook
        wb.save(EXCEL_FILE)
        wb.close()
        print(f"Created sample Excel file: {EXCEL_FILE}")

# Load data from Excel
def load_data():
    if not os.path.exists(EXCEL_FILE):
        create_sample_excel()
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb[SHEET_NAME]
    
    # Get headers from the first row
    headers = [cell.value for cell in sheet[1]]
    
    # Get data from the remaining rows
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, value in enumerate(row):
            # Convert datetime objects to string
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            header = headers[i]
            # Migrate old "Initial Remarks" to "Initial Screening"
            if header == 'Initial Remarks':
                header = 'Initial Screening'
            row_data[header] = str(value) if value is not None else ''
        data.append(row_data)
    
    return data

# Save data to Excel
def save_data(data):
    if not os.path.exists(EXCEL_FILE):
        create_sample_excel()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb[SHEET_NAME]

    # Clear existing data (keep headers)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # Write new data
    headers = [cell.value for cell in sheet[1]]
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=row_num, column=col_num).value = row_data.get(header, '')

    wb.save(EXCEL_FILE)
    wb.close()
    print(f"Data saved to Excel: {data}")
    
    # Desired field order (keep 'Date' at the beginning)
    desired_fields = [
        'Name', 'Email ID', 'Contact Number', 'Interested Position', 'Current Role',
        'Current Organization', 'Current Location', 'Current CTC per Annum',
        'Expected CTC per Annum', 'Total Years of Experience', 'Notice Period',
        'Interview Status', 'Application Status', 'Referred By', 'Comments',
        'In Notice', 'Immediate Joiner', 'Offers in Hand', 'Offered CTC',
        'Location Preference', 'Certifications', 'Resume', 'LinkedIn Profile',
        # Stage-specific remarks that should be persisted
        'Initial Screening', 'Round 1 Remarks', 'Round 2 Remarks',
        # General/legacy remarks
        'Remarks', 'Reject Mail Sent', 'Final Remarks', 'Month Count'
    ]
    
    # Build ordered headers: Date + desired fields present + any remaining headers
    ordered_headers = []
    if 'Date' in headers:
        ordered_headers.append('Date')
    ordered_headers.extend([h for h in desired_fields if h in headers])
    # Include any headers not in desired list (e.g., 'Reference')
    ordered_headers.extend([h for h in headers if h not in ordered_headers])
    
    # If there are desired fields missing from headers, append them so they are created
    ordered_headers.extend([h for h in desired_fields if h not in ordered_headers])
    
    # Rewrite headers in desired order
    for col_num, header in enumerate(ordered_headers, 1):
        sheet.cell(row=1, column=col_num).value = header
    
    # Clear existing data (except headers)
    for row in range(sheet.max_row, 1, -1):
        sheet.delete_rows(row)
    
    # Add updated data
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(ordered_headers, 1):
            # Migrate old "Initial Remarks" to "Initial Screening"
            if header == 'Initial Screening':
                value = row_data.get('Initial Screening') or row_data.get('Initial Remarks', '')
            else:
                value = row_data.get(header, '')
            sheet.cell(row=row_num, column=col_num).value = value
    
    # Save the workbook
    wb.save(EXCEL_FILE)

# Initialize user database
def init_user_db():
    """Initialize the user database with admin user"""
    os.makedirs('instance', exist_ok=True)
    conn = sqlite3.connect(USER_DB)
    cursor = conn.cursor()
    
    # Create users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            is_admin INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Check if admin user exists
    cursor.execute('SELECT COUNT(*) FROM users WHERE username = ?', (ADMIN_USERNAME,))
    if cursor.fetchone()[0] == 0:
        # Create default admin user
        password_hash = hashlib.sha256(ADMIN_PASSWORD.encode()).hexdigest()
        cursor.execute('''
            INSERT INTO users (username, password_hash, is_admin)
            VALUES (?, ?, 1)
        ''', (ADMIN_USERNAME, password_hash))
        conn.commit()
    
    conn.close()

# Hash password
def hash_password(password):
    """Hash a password using SHA256"""
    return hashlib.sha256(password.encode()).hexdigest()

# Verify password
def verify_password(password, password_hash):
    """Verify a password against its hash"""
    return hash_password(password) == password_hash

# Check if user is admin
def is_admin():
    """Check if the current user is an admin"""
    return session.get('is_admin', False)

# Login route
@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # Check against database
        conn = sqlite3.connect(USER_DB)
        cursor = conn.cursor()
        cursor.execute('SELECT password_hash, is_admin FROM users WHERE username = ?', (username,))
        user = cursor.fetchone()
        conn.close()
        
        if user and verify_password(password, user[0]):
            session['logged_in'] = True
            session['username'] = username
            session['is_admin'] = bool(user[1])
            return redirect(url_for('index'))
        else:
            error = 'Invalid credentials. Please try again.'
    
    return render_template('login.html', error=error)

# Logout route
@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    session.pop('username', None)
    session.pop('is_admin', None)
    return redirect(url_for('login'))

# Check if user is logged in
def login_required(f):
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# Check if user is admin (decorator)
def admin_required(f):
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        if not session.get('is_admin'):
            return jsonify({"status": "error", "message": "Admin access required"}), 403
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/api/data', methods=['GET'])
@login_required
def get_data():
    data = load_data()
    is_admin_user = is_admin()  # Check if the user is an admin
    return jsonify({"data": data, "is_admin": is_admin_user})

@app.route('/api/data', methods=['POST'])
@login_required
def add_data():
    try:
        new_data = request.json
        data = load_data()
        data.append(new_data)
        save_data(data)
        return jsonify({"status": "success", "message": "Data added successfully"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/data/<int:index>', methods=['PUT'])
@login_required
def update_data(index):
    try:
        update_data = request.json
        data = load_data()
        
        # Check if index is valid
        if 0 <= index < len(data):
            # Update the data at the specified index
            for key, value in update_data.items():
                # Convert specific fields to appropriate types if necessary
                if key in ['Current CTC per Annum', 'Expected CTC per Annum', 'Offered CTC']:
                    try:
                        data[index][key] = int(value)
                    except (ValueError, TypeError):
                        data[index][key] = value  # Keep original if conversion fails
                else:
                    data[index][key] = value
            
            save_data(data)
            return jsonify({"status": "success", "message": "Data updated successfully"})
        else:
            return jsonify({"status": "error", "message": f"No record found at index {index}"}), 404
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/data/<int:index>', methods=['DELETE'])
@login_required
def delete_data(index):
    try:
        data = load_data()
        
        # Check if index is valid
        if 0 <= index < len(data):
            # Delete the data at the specified index
            del data[index]
            save_data(data)
            return jsonify({"status": "success", "message": "Data deleted successfully"})
        else:
            return jsonify({"status": "error", "message": f"No record found at index {index}"}), 404
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/analysis/summary', methods=['GET'])
@login_required
def get_summary():
    try:
        data = load_data()
        if not data:
            return jsonify({"status": "error", "message": "No data available"}), 404
        
        # Get numerical columns (assuming columns with numeric values)
        numeric_cols = ['Current CTC per Annum', 'Expected CTC per Annum', 'Offered CTC']
        
        summary = {}
        for col in numeric_cols:
            values = [float(item[col]) for item in data if item[col] and item[col].replace('.', '', 1).isdigit()]
            if values:
                summary[col] = {
                    "mean": sum(values) / len(values),
                    "min": min(values),
                    "max": max(values),
                    "count": len(values)
                }
        
        return jsonify(summary)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/analysis/group/<column>', methods=['GET'])
@login_required
def group_analysis(column):
    try:
        data = load_data()
        if not data:
            return jsonify({"status": "error", "message": "No data available"}), 404
        
        # Check if column exists
        if not data or column not in data[0]:
            return jsonify({"status": "error", "message": f"Column {column} not found"}), 400
        
        # Get numerical columns for aggregation
        numeric_cols = ['Current CTC per Annum', 'Expected CTC per Annum', 'Offered CTC']
        
        # Group by the specified column
        grouped_data = {}
        for item in data:
            group_key = item.get(column, 'Unknown')
            if group_key not in grouped_data:
                grouped_data[group_key] = {col: [] for col in numeric_cols}
            
            # Add numeric values to the appropriate group
            for col in numeric_cols:
                if item.get(col) and item[col].replace('.', '', 1).isdigit():
                    grouped_data[group_key][col].append(float(item[col]))
        
        # Calculate averages for each group
        result = []
        for group_key, values in grouped_data.items():
            group_result = {column: group_key}
            for col, numbers in values.items():
                if numbers:
                    group_result[col + '_avg'] = sum(numbers) / len(numbers)
                    group_result[col + '_count'] = len(numbers)
                else:
                    group_result[col + '_avg'] = 0
                    group_result[col + '_count'] = 0
            result.append(group_result)
        
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/dropdown-options', methods=['GET'])
@login_required
def get_dropdown_options():
    """Return all dropdown options for form fields"""
    dropdown_options = {
        'Interested Position': [
            'Backend Developer', 
            'Frontend Developer', 
            'Full Stack Developer', 
            'DevOps Engineer', 
            'Data Engineer', 
            'Data Scientist',
            'UI/UX Designer',
            'Product Manager',
            'QA Engineer'
        ],
        'Current Role': [
            'Software Engineer',
            'Senior Software Engineer',
            'Lead Engineer',
            'Engineering Manager',
            'Architect',
            'QA Engineer',
            'DevOps Engineer',
            'Data Engineer',
            'Data Scientist',
            'Product Manager',
            'UI/UX Designer'
        ],
        'Current Location': [
            'Bangalore', 
            'Chennai', 
            'Hyderabad', 
            'Mumbai', 
            'Delhi', 
            'Pune', 
            'Kolkata',
            'Remote'
        ],
        'Location Preference': [
            'Bangalore', 
            'Chennai', 
            'Hyderabad', 
            'Mumbai', 
            'Delhi', 
            'Pune', 
            'Kolkata',
            'Remote'
        ],
        'Total Years of Experience': [
            '0-1 years',
            '1-2 years',
            '2-3 years',
        ],
        'Notice Period': [
            'Immediate',
            '15 days',
            '30 days',
            '60 days',
            '90 days'
        ],
        'In Notice': ['Yes', 'No'],
        'Immediate Joiner': ['Yes', 'No'],
        'Offers in Hand': ['Yes', 'No'],
        'Interview Status': [
            'Applied',
            'Profile Screening Comp',
            'Voice Screening Comp',
            'Tech Inter Sched',
            'Tech Inter Comp',
            'Code Inter Sched',
            'Code Inter Comp',
            'HR Inter Sched',
            'HR Inter Comp',
            'Offer',
            'Pending Final Noti',
            'References',
            'All Completed'
        ],
        'Application Status': [
            'Proceed Further',
            'On Hold',
            'No Resp Call/Email',
            'Did Not Join',
            'Sent',
            'Recieved',
            'In Notice',
            'Accepted',
            'Rejected',
            'Joined'
        ],
        'Reject Mail Sent': ['Yes', 'No']
    }
    
    return jsonify(dropdown_options)

# User Management Routes (Admin Only)
@app.route('/users')
@login_required
def users_page():
    """User management page (admin only)"""
    if not is_admin():
        return redirect(url_for('index'))
    print(f"isAdmin status before rendering users.html: {is_admin()}")
    return render_template('users.html', isAdmin=is_admin())

@app.route('/api/users', methods=['GET'])
@admin_required
def get_users():
    """Get all users (admin only)"""
    conn = sqlite3.connect(USER_DB)
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, is_admin, created_at FROM users ORDER BY created_at DESC')
    users = cursor.fetchall()
    conn.close()
    
    users_list = []
    for user in users:
        users_list.append({
            'id': user[0],
            'username': user[1],
            'is_admin': bool(user[2]),
            'created_at': user[3]
        })
    
    return jsonify(users_list)

@app.route('/api/users', methods=['POST'])
@admin_required
def add_user():
    """Add a new user (admin only)"""
    try:
        data = request.json
        username = data.get('username')
        password = data.get('password')
        is_admin_flag = data.get('is_admin', False)
        
        if not username or not password:
            return jsonify({"status": "error", "message": "Username and password are required"}), 400
        
        conn = sqlite3.connect(USER_DB)
        cursor = conn.cursor()
        
        # Check if username already exists
        cursor.execute('SELECT COUNT(*) FROM users WHERE username = ?', (username,))
        if cursor.fetchone()[0] > 0:
            conn.close()
            return jsonify({"status": "error", "message": "Username already exists"}), 400
        
        # Add user
        password_hash = hash_password(password)
        cursor.execute('''
            INSERT INTO users (username, password_hash, is_admin)
            VALUES (?, ?, ?)
        ''', (username, password_hash, 1 if is_admin_flag else 0))
        conn.commit()
        conn.close()
        
        return jsonify({"status": "success", "message": "User added successfully"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    """Delete a user (admin only)"""
    try:
        # Prevent deleting yourself
        if user_id == session.get('user_id'):
            return jsonify({"status": "error", "message": "Cannot delete your own account"}), 400
        
        conn = sqlite3.connect(USER_DB)
        cursor = conn.cursor()
        
        # Check if user exists
        cursor.execute('SELECT username FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()
        if not user:
            conn.close()
            return jsonify({"status": "error", "message": "User not found"}), 404
        
        # Prevent deleting admin user
        cursor.execute('SELECT is_admin FROM users WHERE id = ?', (user_id,))
        if cursor.fetchone()[0] == 1:
            # Check if there are other admins
            cursor.execute('SELECT COUNT(*) FROM users WHERE is_admin = 1 AND id != ?', (user_id,))
            if cursor.fetchone()[0] == 0:
                conn.close()
                return jsonify({"status": "error", "message": "Cannot delete the last admin user"}), 400
        
        # Delete user
        cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()
        conn.close()
        
        return jsonify({"status": "success", "message": "User deleted successfully"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == '__main__':
    # Initialize user database
    init_user_db()
    # Apply header ordering to existing Excel data on startup
    try:
        existing_data = load_data()
        save_data(existing_data)
    except Exception:
        pass
    app.run(debug=True)